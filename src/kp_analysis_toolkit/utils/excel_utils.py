"""
Common Excel utilities for the analysis toolkit.

Provides reusable Excel formatting, table creation, and export functionality
that can be used across all modules in the toolkit.

This module provides backward compatibility for existing code while delegating
to the new DI-based Excel export services when available.
"""

from __future__ import annotations

import random
import re
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

from kp_analysis_toolkit.utils.di_state import create_excel_export_di_manager

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from kp_analysis_toolkit.core.services.excel_export.service import (
        ExcelExportService,
    )

# Global DI state manager for Excel export services
_di_manager, _get_service, _set_service, _clear_service = (
    create_excel_export_di_manager()
)


def _get_excel_export_service() -> object | None:
    """Get the Excel export service if DI is available."""
    return _get_service()


def _set_excel_export_service(service: object) -> None:
    """Set the Excel export service for DI integration."""
    _set_service(service)


# Backward compatibility functions that delegate to DI services when available


def sanitize_sheet_name(name: str) -> str:
    """
    Sanitize a string to be used as an Excel sheet name.

    This function supports dependency injection when available, falling back to
    direct implementation for backward compatibility.

    Args:
        name: The string to sanitize

    Returns:
        A string safe to use as an Excel sheet name

    """
    # Try to use DI-based Excel export service first
    excel_service = _get_excel_export_service()
    if excel_service is not None:
        try:
            return excel_service.sheet_name_sanitizer.sanitize_sheet_name(name)  # type: ignore[attr-defined]
        except (AttributeError, Exception):  # noqa: S110
            # Fall back to direct implementation if DI fails
            pass

    # Direct implementation fallback
    if not name:
        return "Unnamed_Sheet"

    # Remove invalid characters and replace with underscores
    sanitized: str = re.sub(r"[\\/*\[\]:?]", "_", name)

    # Remove extra spaces and replace with underscores
    sanitized = re.sub(r"\s+", "_", sanitized)

    # Remove leading/trailing underscores
    sanitized = sanitized.strip("_")

    # Ensure we have something
    if not sanitized:
        sanitized = "Unnamed_Sheet"

    # Limit to 31 characters (Excel limitation)
    max_sheet_name_length = 31
    if len(sanitized) > max_sheet_name_length:
        sanitized = sanitized[: max_sheet_name_length - 3] + "..."

    return sanitized


def get_column_letter(col_num: int) -> str:
    """
    Convert column number to Excel column letter.

    This function supports dependency injection when available, falling back to
    direct implementation for backward compatibility.

    Args:
        col_num: Column number (1-indexed)

    Returns:
        Excel column letter(s)

    """
    # Try to use DI-based Excel export service first
    excel_service = _get_excel_export_service()
    if excel_service is not None:
        try:
            return excel_service.sheet_name_sanitizer.get_column_letter(col_num)  # type: ignore[attr-defined]
        except (AttributeError, Exception):  # noqa: S110
            # Fall back to direct implementation if DI fails
            pass

    # Direct implementation fallback
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(col_num % 26 + ord("A")) + result
        col_num //= 26
    return result


def auto_adjust_column_widths(worksheet: Worksheet, df: pd.DataFrame) -> None:
    """
    Auto-adjust column widths based on content.

    This function supports dependency injection when available, falling back to
    direct implementation for backward compatibility.

    Args:
        worksheet: The worksheet to adjust
        df: DataFrame that was written to the worksheet

    """
    # Try to use DI-based Excel export service first
    excel_service = _get_excel_export_service()
    if excel_service is not None:
        try:
            excel_service.column_width_adjuster.auto_adjust_column_widths(  # type: ignore[attr-defined]
                worksheet,
                df,
            )
        except (AttributeError, Exception):  # noqa: S110
            # Fall back to direct implementation if DI fails
            pass
        else:
            return

    # Direct implementation fallback
    for col_num, column_name in enumerate(df.columns, 1):
        column_letter = get_column_letter(col_num)

        # Calculate max width needed
        max_length = len(str(column_name))  # Header length

        # Check data in the column
        for value in df.iloc[:, col_num - 1]:
            if pd.notna(value):
                length = len(str(value))
                max_length = max(max_length, length)

        # Set a reasonable width (with limits)
        width = min(max(max_length + 2, 10), 50)
        worksheet.column_dimensions[column_letter].width = width


def format_date_columns(  # noqa: C901, PLR0912
    worksheet: Worksheet,
    df: pd.DataFrame,
    startrow: int = 1,
) -> None:
    """
    Format date columns in the worksheet.

    This function supports dependency injection when available, falling back to
    direct implementation for backward compatibility.

    Args:
        worksheet: The worksheet to format
        df: DataFrame that was written to the worksheet
        startrow: Row where the data starts (1-indexed)

    """
    # Try to use DI-based Excel export service first
    excel_service = _get_excel_export_service()
    if excel_service is not None:
        try:
            excel_service.date_formatter.format_date_columns(  # type: ignore[attr-defined]
                worksheet,
                df,
                startrow,
            )
        except (AttributeError, Exception):  # noqa: S110
            # Fall back to direct implementation if DI fails
            pass
        else:
            return

    # Direct implementation fallback
    for col_num, column_name in enumerate(df.columns, 1):
        column_letter = get_column_letter(col_num)

        # Check if this looks like a date column
        is_date_column = False
        if "date" in str(column_name).lower() or "time" in str(column_name).lower():
            is_date_column = True
        else:
            # Sample some values to see if they look like dates
            sample_values = df.iloc[:5, col_num - 1].dropna()
            date_count = 0
            for value in sample_values:
                if pd.notna(value):
                    str_value = str(value)
                    # Simple date pattern check
                    if re.match(r"\d{4}-\d{2}-\d{2}", str_value) or re.match(
                        r"\d{2}/\d{2}/\d{4}",
                        str_value,
                    ):
                        date_count += 1
            if date_count >= len(sample_values) * 0.5:  # 50% or more look like dates
                is_date_column = True

        if is_date_column:
            # Apply date formatting to the entire column
            for row_num in range(startrow + 1, startrow + len(df) + 1):
                cell = f"{column_letter}{row_num}"

                # Try to parse and standardize date values
                try:
                    value = worksheet[cell].value
                    if value and pd.notna(value):
                        # Try to parse various date formats
                        parsed_date = None
                        str_value = str(value)

                        # Common date patterns
                        patterns: list[str] = [
                            "%Y-%m-%d",
                            "%m/%d/%Y",
                            "%d/%m/%Y",
                            "%Y-%m-%d %H:%M:%S",
                            "%m/%d/%Y %H:%M:%S",
                        ]

                        for pattern in patterns:
                            try:
                                parsed_date = datetime.strptime(  # noqa: DTZ007
                                    str_value,
                                    pattern,
                                )
                                break
                            except ValueError:
                                continue

                        # If we successfully parsed the date, standardize it
                        if parsed_date:
                            # Update the cell with the standardized date string
                            worksheet[cell].value = parsed_date.strftime("%Y-%m-%d")
                except Exception:  # noqa: BLE001, S110
                    # If parsing fails, keep the original value
                    pass

                # Apply Excel date formatting
                worksheet[cell].number_format = "yyyy-mm-dd"


def set_table_alignment(worksheet: Worksheet, table_range: str) -> None:
    """
    Set alignment for table cells.

    This function supports dependency injection when available, falling back to
    direct implementation for backward compatibility.

    Args:
        worksheet: The worksheet containing the table
        table_range: Excel range string (e.g., "A1:D10")

    """
    # Try to use DI-based Excel export service first
    excel_service = _get_excel_export_service()
    if excel_service is not None:
        try:
            excel_service.excel_formatter.set_table_alignment(  # type: ignore[attr-defined]
                worksheet,
                table_range,
            )
        except (AttributeError, Exception):  # noqa: S110
            # Fall back to direct implementation if DI fails
            pass
        else:
            return

    # Direct implementation fallback
    for row in worksheet[table_range]:
        for cell in row:
            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )


def adjust_row_heights(
    worksheet: Worksheet,
    df: pd.DataFrame,
    startrow: int = 1,
) -> None:
    """
    Adjust row heights based on content.

    This function supports dependency injection when available, falling back to
    direct implementation for backward compatibility.

    Args:
        worksheet: The worksheet to adjust
        df: DataFrame that was written to the worksheet
        startrow: Row where the data starts (1-indexed)

    """
    # Try to use DI-based Excel export service first
    excel_service = _get_excel_export_service()
    if excel_service is not None:
        try:
            excel_service.row_height_adjuster.adjust_row_heights(  # type: ignore[attr-defined]
                worksheet,
                df,
                startrow,
            )
        except (AttributeError, Exception):  # noqa: S110
            # Fall back to direct implementation if DI fails
            pass
        else:
            return

    # Direct implementation fallback
    # Set minimum row height for header
    worksheet.row_dimensions[startrow].height = 20

    # Adjust data row heights
    for row_num in range(startrow + 1, startrow + len(df) + 1):
        max_lines = 1

        # Check each cell in the row for line breaks
        for col_num in range(1, len(df.columns) + 1):
            column_letter = get_column_letter(col_num)
            cell_value = worksheet[f"{column_letter}{row_num}"].value

            if cell_value and isinstance(cell_value, str):
                lines = cell_value.count("\n") + 1
                max_lines = max(max_lines, lines)

        # Set row height (approximate 15 points per line)
        height = max(15, min(max_lines * 15, 100))
        worksheet.row_dimensions[row_num].height = height


def format_as_excel_table(
    worksheet: Worksheet,
    df: pd.DataFrame,
    startrow: int = 1,
) -> None:
    """
    Format DataFrame as an Excel table with proper styling.

    This function supports dependency injection when available, falling back to
    direct implementation for backward compatibility.

    Args:
        worksheet: The worksheet to format
        df: DataFrame that was written to the worksheet
        startrow: Row where the data starts (1-indexed)

    """
    # Try to use DI-based Excel export service first
    excel_service = _get_excel_export_service()
    if excel_service is not None:
        try:
            excel_service.table_generator.format_as_excel_table(  # type: ignore[attr-defined]
                worksheet,
                df,
                startrow,
            )
        except (AttributeError, Exception):  # noqa: S110
            # Fall back to direct implementation if DI fails
            pass
        else:
            return

    # Direct implementation fallback
    if df.empty:
        return

    # Calculate table range
    end_row: int = startrow + len(df)
    end_col_letter: str = get_column_letter(len(df.columns))
    table_range: str = f"A{startrow}:{end_col_letter}{end_row}"

    # Create unique table name
    random_suffix: int = random.randint(1000, 9999)
    table_name: str = f"Table_{worksheet.title}_{random_suffix}".replace(
        " ",
        "_",
    ).replace(
        "-",
        "_",
    )
    # Ensure table name doesn't start with a number or contain invalid chars
    table_name = "".join(c if c.isalnum() or c == "_" else "_" for c in table_name)
    if table_name[0].isdigit():
        table_name = "T_" + table_name

    # Create table
    table = Table(displayName=table_name, ref=table_range)

    # Add table style
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style

    # Add table to worksheet
    try:
        worksheet.add_table(table)
    except ValueError as e:
        message = f"Error adding table to worksheet '{worksheet.title}': {e}. "
        raise ValueError(message) from e

    # Set top-left alignment for all cells in table range
    set_table_alignment(worksheet, table_range)

    # Format date columns
    format_date_columns(worksheet, df, startrow)

    # Auto-adjust column widths
    auto_adjust_column_widths(worksheet, df)

    # Adjust row heights based on content
    adjust_row_heights(worksheet, df, startrow)


def export_dataframe_to_excel(
    df: pd.DataFrame,
    output_path: Path,
    sheet_name: str = "Sheet1",
    title: str | None = None,
    *,
    as_table: bool = True,
) -> None:
    """
    General-purpose DataFrame to Excel export with formatting.

    This function supports dependency injection when available, falling back to
    direct implementation for backward compatibility.

    Args:
        df: DataFrame to export
        output_path: Path to save Excel file
        sheet_name: Name of the worksheet
        title: Optional title to add at the top
        as_table: Whether to format as Excel table

    """
    # Try to use DI-based Excel export service first
    excel_service = _get_excel_export_service()
    if excel_service is not None:
        try:
            excel_service.export_dataframe_to_excel(  # type: ignore[attr-defined]
                df,
                output_path,
                sheet_name,
                title,
                as_table=as_table,
            )
        except (AttributeError, Exception):  # noqa: S110
            # Fall back to direct implementation if DI fails
            pass
        else:
            return

    # Direct implementation fallback
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Sanitize sheet name for Excel compatibility
    sanitized_sheet_name = sanitize_sheet_name(sheet_name)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        start_row = 1 if title else 0
        df.to_excel(
            writer, sheet_name=sanitized_sheet_name, index=False, startrow=start_row,
        )
        worksheet = writer.sheets[sanitized_sheet_name]

        if title:
            worksheet["A1"] = title
            worksheet["A1"].font = Font(bold=True, size=12, color="1F497D")

        if as_table and not df.empty:
            format_as_excel_table(worksheet, df, startrow=start_row + 1)
        else:
            auto_adjust_column_widths(worksheet, df)


# DI Integration Functions


def set_excel_export_service(service: ExcelExportService) -> None:
    """
    Set the Excel export service for dependency injection integration.

    This function allows external code to configure the Excel export service
    that will be used by all backward compatibility functions.

    Args:
        service: The ExcelExportService instance to use

    """
    _set_excel_export_service(service)


def get_excel_export_service() -> object | None:
    """
    Get the current Excel export service if available.

    Returns:
        The current ExcelExportService instance or None if not configured

    """
    return _get_excel_export_service()


def clear_excel_export_service() -> None:
    """
    Clear the Excel export service, forcing fallback to direct implementations.

    This function is useful for testing or when you want to ensure
    direct implementations are used instead of DI services.

    """
    _clear_service()


def get_di_state() -> object:
    """
    Get the DI state for compatibility with test expectations.

    Returns:
        The DI state manager instance

    """
    return _di_manager
