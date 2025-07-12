"""Formatting utilities for Excel export."""

from __future__ import annotations

from datetime import datetime
from typing import TYPE_CHECKING, Any

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

from kp_analysis_toolkit.core.services.excel_export.protocols import (
    ColumnWidthAdjuster,
    DateFormatter,
    ExcelFormatter,
    RowHeightAdjuster,
    TableStyler,
    TitleFormatter,
)
from kp_analysis_toolkit.core.services.excel_export.sheet_management import (
    DefaultSheetNameSanitizer,
)

if TYPE_CHECKING:
    from openpyxl.cell.cell import Cell
    from openpyxl.worksheet.worksheet import Worksheet

    from kp_analysis_toolkit.models.types import DisplayableValue


class DefaultColumnWidthAdjuster(ColumnWidthAdjuster):
    """Default implementation for auto-adjusting column widths."""

    def auto_adjust_column_widths(
        self,
        worksheet: Worksheet,
        data_frame: pd.DataFrame,
    ) -> None:
        """
        Auto-adjust the column widths based on the content of the DataFrame.

        Args:
            worksheet: The worksheet to adjust
            data_frame: The DataFrame containing the data

        """
        for col_num, column_name in enumerate(data_frame.columns, 1):
            column_letter: str = DefaultSheetNameSanitizer().get_column_letter(col_num)
            max_length: int = len(str(column_name))
            for value in data_frame.iloc[:, col_num - 1]:
                # Check if value is not NaN for floats, or not None for other types
                if (isinstance(value, float) and not pd.isna(value)) or (
                    not isinstance(value, float) and value is not None
                ):
                    length: int = len(str(value))
                    max_length = max(max_length, length)
            width: int = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = width


class DefaultDateFormatter(DateFormatter):
    """Default implementation for formatting date columns."""

    def __init__(self) -> None:
        self.DATE_PATTERNS: list[str] = [
            "%Y-%m-%d",
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%Y-%m-%d %H:%M:%S",
            "%m/%d/%Y %H:%M:%S",
        ]

    def _is_date_column(self, column_name: str, sample_values: pd.Series) -> bool:
        """
        Check if a column is likely to contain date values.

        Args:
            column_name: Name of the column
            sample_values: Sample values from the column

        Returns:
            bool: True if the column is likely a date column, False otherwise

        """
        if "date" in column_name.lower() or "time" in column_name.lower():
            return True
        date_count = 0
        for value in sample_values:
            if self._is_date_value(value):
                date_count += 1
        return len(sample_values) > 0 and date_count >= len(sample_values) * 0.5

    def _is_date_value(self, value: DisplayableValue) -> bool:
        """
        Check if a value is a date.

        Args:
            value: The value to check

        Returns:
            bool: True if the value is a date, False otherwise

        """
        return (
            self._is_valid_cell_value(value)
            and self._parse_date(str(value)) is not None
        )

    def _parse_date(self, value: str) -> datetime | None:
        """
        Parse a string value into a datetime object.

        Args:
            value: The string value to parse
        Returns:
            datetime | None: Parsed datetime object or None if parsing fails

        """
        for pattern in self.DATE_PATTERNS:
            try:
                return datetime.strptime(value, pattern)  # noqa: DTZ007
            except ValueError:
                continue
        return None

    def _is_valid_cell_value(self, value: DisplayableValue) -> bool:
        """
        Check if a cell value is valid for date parsing.

        Args:
            value: The cell value to check

        Returns:
            bool: True if the value is valid, False otherwise

        """
        if value is None:
            return False
        # Check for NaN values (primarily for float types)
        if isinstance(value, float):
            return not pd.isna(value)
        # For other types, assume valid if not None
        return True

    def format_date_columns(
        self,
        worksheet: Worksheet,
        data_frame: pd.DataFrame,
        startrow: int = 1,
    ) -> None:
        """
        Format date columns in the worksheet.

        Args:
            worksheet: The worksheet to format
            data_frame: The DataFrame containing the data
            startrow: The starting row for formatting (default: 1)

        """
        for col_num, column_name in enumerate(data_frame.columns, 1):
            column_letter: str = DefaultSheetNameSanitizer().get_column_letter(col_num)
            sample_values: pd.Series[Any] = data_frame.iloc[:5, col_num - 1].dropna()
            if self._is_date_column(str(column_name), sample_values):
                for row_num in range(startrow + 1, startrow + len(data_frame) + 1):
                    cell: str = f"{column_letter}{row_num}"
                    try:
                        value: DisplayableValue = worksheet[cell].value
                        if self._is_valid_cell_value(value):
                            parsed_date: datetime | None = self._parse_date(str(value))
                            if parsed_date:
                                worksheet[cell].value = parsed_date.strftime("%Y-%m-%d")

                                # Apply Excel date formatting
                                worksheet[cell].number_format = "yyyy-mm-dd"
                    except Exception:  # noqa: BLE001, S110
                        # Ignore errors and keep the original value
                        pass


class DefaultRowHeightAdjuster(RowHeightAdjuster):
    """Default implementation for adjusting row heights."""

    def adjust_row_heights(
        self,
        worksheet: Worksheet,
        data_frame: pd.DataFrame,
        startrow: int = 1,
    ) -> None:
        """
        Adjust the row heights in the worksheet based on the content.

        Args:
            worksheet: The worksheet to adjust
            data_frame: The DataFrame containing the data
            startrow: The starting row for adjustment (default: 1)

        """
        worksheet.row_dimensions[startrow].height = 20
        for row_num in range(startrow + 1, startrow + len(data_frame) + 1):
            max_lines = 1
            for col_num in range(1, len(data_frame.columns) + 1):
                column_letter: str = DefaultSheetNameSanitizer().get_column_letter(
                    col_num,
                )
                cell_value: DisplayableValue = worksheet[
                    f"{column_letter}{row_num}"
                ].value
                if cell_value and isinstance(cell_value, str):
                    lines: int = cell_value.count("\n") + 1
                    max_lines = max(max_lines, lines)
            height: float = max(15, min(max_lines * 15, 100))
            worksheet.row_dimensions[row_num].height = height


class DefaultExcelFormatter(ExcelFormatter):
    """Default implementation for Excel table alignment and column width."""

    def __init__(
        self,
        column_width_adjuster: ColumnWidthAdjuster | None = None,
    ) -> None:
        """
        Initialize the Excel formatter.

        Args:
            column_width_adjuster: The column width adjuster to use.
                                 If None, uses DefaultColumnWidthAdjuster.

        """
        self._column_width_adjuster: (
            ColumnWidthAdjuster | DefaultColumnWidthAdjuster
        ) = column_width_adjuster or DefaultColumnWidthAdjuster()

    def set_table_alignment(self, worksheet: Worksheet, table_range: str) -> None:
        """
        Set the alignment for all cells in the specified table range.

        Args:
            worksheet: The worksheet containing the table
            table_range: The range of the table (e.g., "A1:D10")

        """
        for row in worksheet[table_range]:
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True,
                )

    def auto_adjust_column_widths(
        self,
        worksheet: Worksheet,
        data_frame: pd.DataFrame,
    ) -> None:
        """
        Auto-adjust the column widths based on the content of the DataFrame.

        Args:
            worksheet: The worksheet to adjust
            data_frame: The DataFrame containing the data

        """
        self._column_width_adjuster.auto_adjust_column_widths(worksheet, data_frame)


class DefaultTableStyler(TableStyler):
    def apply_style(self, table: Table) -> None:
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        """
        Apply the table style to the given table.
        Args:
            table: The table to style

        """
        table.tableStyleInfo = style


class DefaultTitleFormatter(TitleFormatter):
    def apply_title_format(
        self,
        worksheet: Worksheet,
        title: str,
        row: int = 1,
        col: int = 1,
    ) -> None:
        """
        Apply formatting to the title cell.

        Args:
            worksheet: The worksheet to format
            title: The title text
            row: The row number for the title (default: 1)
            col: The column number for the title (default: 1)

        """
        cell: Cell = worksheet.cell(row=row, column=col, value=title)
        cell.font = Font(bold=True, size=12, color="1F497D")
